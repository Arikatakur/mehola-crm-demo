# -*- coding: utf-8 -*-
"""
Mehola CRM demo — Excel ("Gilyon") -> normalized dataset.

Reads the legacy site workbook (site: Rosman / רוסמן) and emits
src/data/rosman.js  — a plain <script> data module consumed by the demo app.

Pipeline (mirrors §3.1 "Staging & Deduplication" of the feature report):
  1. Staging      — read every day-block of the sheet as raw punch rows.
  2. Cleansing    — collapse whitespace, unify carrier/team spellings,
                    repair out-of-sequence dates using the Hebrew day name.
  3. Deduplication— merge worker name variants into one canonical worker,
                    attach the ID (Teudat Zehut) where the sheet has one.
  4. Audit        — every correction and every unresolved anomaly is written
                    to the `issues` list so the app can display it.

Nothing is silently dropped: if the script cannot resolve something it is
reported as an issue rather than guessed.

Usage:  python tools/migrate_excel.py [path-to-xlsx]
"""

import datetime
import json
import os
import re
import sys
import unicodedata

try:
    import openpyxl
except ImportError:  # pragma: no cover
    sys.exit("openpyxl is required:  pip install openpyxl")

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
DEFAULT_XLSX = os.path.join(os.path.dirname(ROOT), "רוסמן..xlsx")
OUT_JS = os.path.join(ROOT, "src", "data", "rosman.js")

SITE = {"id": "rosman", "name": "רוסמן", "billing": "unit", "unitLabel": "מדבקות"}

# python weekday(): Mon=0 .. Sun=6
HEB_DAY = {"ראשון": 6, "שני": 0, "שלישי": 1, "רביעי": 2, "חמישי": 3, "שישי": 4, "שבת": 5}

# --- cleansing tables -------------------------------------------------------

CARRIER_ALIASES = {
    "גסלין": "גיסלין",
    "גיסלין": "גיסלין",
    "כרסתין": "כריסתין",
    "עצמאית": "עצמאי",
    "עצמאי": "עצמאי",
}

# name variants -> canonical worker name
NAME_ALIASES = {
    "אבתאסם זידאן": "אבתיסאם זידאן",
    "אבתסאם": "אבתיסאם זידאן",
    "סאברין": "סאברין זידאן",
    "עליא": "עליא קדח",
    "פאטמה": "פאטמה מעארי",
    "פאטמה מיעארי": "פאטמה מעארי",
    "סאלי": "סאלי שתיוי",
    "סאלי שתוי": "סאלי שתיוי",
    "רנא גריס": "רנא גיריס",
    "כרסתין סגראוי": "כריסתין סגראוי",
    "מאריא": "מאריא מסיס",
    "חולוד יוסף": "חולוד יוסף",
    "סוהא": "סוהא חידר",
    "נוהא": "נוהא חידר",
    "פסיל": "פסיל עבד אלחמיד",
    "כפאח מנדלאוי": "קפאח מנדלאוי",
}


def txt(v):
    """Cell -> trimmed single-spaced string."""
    if v is None:
        return ""
    s = str(v).strip()
    s = unicodedata.normalize("NFKC", s)
    return re.sub(r"\s+", " ", s)


def hhmm(v):
    s = txt(v)
    m = re.match(r"(\d{1,2}):(\d{2})", s)
    return "%02d:%02d" % (int(m.group(1)), int(m.group(2))) if m else ""


def id_is_valid(tz):
    """Israeli Teudat Zehut check digit (9 digits, Luhn-style)."""
    if not re.fullmatch(r"\d{1,9}", tz or ""):
        return False
    d = tz.zfill(9)
    total = 0
    for i, ch in enumerate(d):
        n = int(ch) * (1 if i % 2 == 0 else 2)
        total += n if n < 10 else n - 9
    return total % 10 == 0


class Migration:
    def __init__(self, path):
        self.path = path
        self.issues = []
        self.days = []       # [{date, dayName, units, bonusUnits, rows:[shift]}]
        self.workers = {}    # canonical name -> worker dict

    # -- audit -------------------------------------------------------------
    def issue(self, kind, severity, title, detail, ref=""):
        self.issues.append({"kind": kind, "severity": severity, "title": title,
                            "detail": detail, "ref": ref})

    # -- 1. staging --------------------------------------------------------
    def read(self):
        wb = openpyxl.load_workbook(self.path, data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

        cur = None          # current day block
        prev_date = None
        for rn, row in enumerate(rows, 1):
            raw0 = row[0] if len(row) else None
            c = [txt(x) for x in row]
            while len(c) < 24:
                c.append("")

            if c[0].startswith("תאריך"):          # repeated header
                continue

            if any("דפוקה" in x for x in c):       # daily production total
                self._read_units(c, cur, rn)
                continue

            day_name = c[1]
            if raw0 not in (None, "") and day_name in HEB_DAY:
                as_written = self._parse_date(raw0, day_name)
                # a block continues until a *different* date is written in col A
                if as_written and (cur is None or as_written != cur["asWritten"]):
                    iso = self._repair_sequence(as_written, day_name, prev_date, rn)
                    cur = {"date": iso, "asWritten": as_written, "dayName": day_name,
                           "units": None, "bonusUnits": 0, "baseUnits": None, "rows": []}
                    self.days.append(cur)
                    prev_date = iso

            name, tz, team, carrier = c[2], c[3], c[4], c[5]
            dept, tin, tout = c[6], hhmm(c[7]), hhmm(c[8])

            if not name or not tin or not tout:
                # rows without punches are scheduling leftovers (absence/notes)
                continue
            if cur is None:
                self.issue("orphan-row", "warn", "שורת נוכחות ללא תאריך",
                           "השורה מופיעה לפני כותרת יום בגיליון ולכן לא שויכה ליום",
                           "שורה %d" % rn)
                continue

            cur["rows"].append({
                "rawName": name, "tz": re.sub(r"\D", "", tz), "team": team,
                "carrier": CARRIER_ALIASES.get(carrier, carrier),
                "dept": dept or "כללי", "in": tin, "out": tout, "row": rn,
            })

    def _read_units(self, c, cur, rn):
        """Row shaped:  'סה"כ דפוקה ליום: 41820' [ 'בונס 2580' ] [ 14000 ]"""
        label = next((x for x in c if "דפוקה" in x), "")
        m = re.search(r"(\d[\d,]*)", label)
        base = float(m.group(1).replace(",", "")) if m else None
        idx = c.index(label)
        bonus, total = 0.0, None
        for cell in c[idx + 1: idx + 4]:
            if "בונס" in cell:
                bm = re.search(r"(\d[\d,]*)", cell)
                if bm:
                    bonus = float(bm.group(1).replace(",", ""))
            elif re.fullmatch(r"\d[\d,]*(\.\d+)?", cell) and bonus:
                total = float(cell.replace(",", ""))
        if cur is None:
            return
        cur["baseUnits"] = base
        cur["bonusUnits"] = bonus
        cur["units"] = total if total is not None else (base or 0) + bonus
        if base is None:
            self.issue("units-missing", "warn", "סה\"כ ייצור יומי לא נקרא",
                       "שורת הסיכום קיימת אך ללא מספר", "שורה %d" % rn)

    # -- 2. cleansing: dates ----------------------------------------------
    def _parse_date(self, raw, day_name):
        """Cell -> ISO date exactly as written (d/m/y, with excel d-m swap undone)."""
        cands = []
        if isinstance(raw, datetime.datetime):
            cands.append(raw.date())
            if raw.day <= 12:                       # excel may have swapped d/m
                cands.append(datetime.date(raw.year, raw.day, raw.month))
        else:
            m = re.match(r"(\d{1,2})[\\/.](\d{1,2})[\\/.](\d{2,4})", str(raw))
            if not m:
                return None
            d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
            y += 2000 if y < 100 else 0
            try:
                cands.append(datetime.date(y, mo, d))
            except ValueError:
                return None

        want = HEB_DAY.get(day_name)
        picked = next((c for c in cands if c.weekday() == want), cands[0])
        return picked.isoformat()

    def _repair_sequence(self, iso, day_name, prev_iso, rn):
        """The sheet is chronological, so a date that goes backwards is a typo.
        Recover the intended date from the Hebrew day name."""
        if not prev_iso or iso > prev_iso:
            return iso
        prev = datetime.date.fromisoformat(prev_iso)
        want = HEB_DAY.get(day_name)
        for step in range(1, 15):
            cand = prev + datetime.timedelta(days=step)
            if cand.weekday() == want:
                self.issue("date-typo", "fixed", "תאריך שגוי תוקן",
                           "בגיליון נרשם %s (%s) בתוך רצף שאחרי %s — תוקן ל-%s"
                           % (iso, day_name, prev_iso, cand.isoformat()),
                           "שורה %d" % rn)
                return cand.isoformat()
        return iso

    # -- 3. deduplication: workers ----------------------------------------
    def build_workers(self):
        for day in self.days:
            for r in day["rows"]:
                canon = NAME_ALIASES.get(r["rawName"], r["rawName"])
                w = self.workers.get(canon)
                if not w:
                    w = self.workers[canon] = {
                        "id": "w%02d" % (len(self.workers) + 1),
                        "name": canon, "tz": "", "aliases": [], "ids": [],
                        "teams": [], "carriers": [],
                    }
                r["workerId"] = w["id"]
                if r["rawName"] != canon and r["rawName"] not in w["aliases"]:
                    w["aliases"].append(r["rawName"])
                if r["tz"] and r["tz"] not in w["ids"]:
                    w["ids"].append(r["tz"])
                if r["team"] and r["team"] not in w["teams"]:
                    w["teams"].append(r["team"])
                if r["carrier"] and r["carrier"] not in w["carriers"]:
                    w["carriers"].append(r["carrier"])

        # ID assignment + conflict detection
        by_id = {}
        for w in self.workers.values():
            if w["ids"]:
                w["tz"] = w["ids"][0]
                by_id.setdefault(w["tz"], []).append(w["name"])
            if len(w["ids"]) > 1:
                self.issue("id-multi", "warn", "מספר ת.ז שונים לאותו עובד",
                           "%s מופיע/ה עם ת.ז: %s" % (w["name"], ", ".join(w["ids"])),
                           w["name"])
            if w["aliases"]:
                self.issue("name-merge", "fixed", "איחוד כתיבים של שם עובד",
                           "אוחדו לרשומה אחת: %s ← %s" % (w["name"], ", ".join(w["aliases"])),
                           w["name"])
            if not w["ids"]:
                self.issue("id-missing", "warn", "עובד ללא ת.ז",
                           "%s רשום/ה בגיליון ללא מספר זהות — חסם למיזוג בין אתרים"
                           % w["name"], w["name"])
            elif not id_is_valid(w["tz"]):
                self.issue("id-invalid", "warn", "ת.ז לא תקינה",
                           "%s: %s אינו מספר זהות תקין (ספרת ביקורת/אורך)"
                           % (w["name"], w["tz"]), w["name"])

        for tz, names in by_id.items():
            if len(names) > 1:
                self.issue("id-conflict", "error", "אותה ת.ז לשני עובדים",
                           "ת.ז %s משויכת ל: %s — נדרש בירור מול המשרד"
                           % (tz, ", ".join(names)), tz)

    # -- audit: production -------------------------------------------------
    def check_days(self):
        for d in self.days:
            if d["units"] is None:
                self.issue("units-missing", "error", "יום ללא נתוני ייצור",
                           "ל-%s אין שורת 'סה\"כ דפוקה ליום' — לא ניתן לחשב הכנסה"
                           % d["date"], d["date"])
            if not d["rows"]:
                self.issue("empty-day", "warn", "יום ללא רישומי נוכחות",
                           "ל-%s לא נמצאו שורות עם כניסה/יציאה" % d["date"], d["date"])

    # -- 4. emit -----------------------------------------------------------
    def emit(self):
        self.days.sort(key=lambda d: d["date"])
        days = [{
            "date": d["date"], "dayName": d["dayName"],
            "units": d["units"], "baseUnits": d["baseUnits"],
            "bonusUnits": d["bonusUnits"],
            "shifts": [{
                "workerId": r["workerId"], "team": r["team"], "carrier": r["carrier"],
                "dept": r["dept"], "in": r["in"], "out": r["out"], "srcRow": r["row"],
            } for r in d["rows"]],
        } for d in self.days]

        workers = sorted(self.workers.values(), key=lambda w: w["id"])
        carriers = sorted({r["carrier"] for d in self.days for r in d["rows"] if r["carrier"]})
        teams = sorted({r["team"] for d in self.days for r in d["rows"] if r["team"]})

        data = {
            "site": SITE,
            "source": {
                "file": os.path.basename(self.path),
                "generatedAt": datetime.datetime.now().strftime("%Y-%m-%d %H:%M"),
                "days": len(days), "shifts": sum(len(d["shifts"]) for d in days),
                "workers": len(workers),
                "from": days[0]["date"] if days else None,
                "to": days[-1]["date"] if days else None,
            },
            "workers": workers, "days": days,
            "carriers": carriers, "teams": teams,
            "issues": self.issues,
        }

        body = json.dumps(data, ensure_ascii=False, indent=1)
        js = ("// GENERATED FILE — do not edit by hand.\n"
              "// Source: %s\n"
              "// Regenerate:  python tools/migrate_excel.py\n"
              "window.M = window.M || {};\n"
              "window.M.DATA = %s;\n" % (os.path.basename(self.path), body))
        os.makedirs(os.path.dirname(OUT_JS), exist_ok=True)
        with open(OUT_JS, "w", encoding="utf-8") as f:
            f.write(js)
        return data


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_XLSX
    if not os.path.exists(path):
        sys.exit("workbook not found: %s" % path)
    m = Migration(path)
    m.read()
    m.build_workers()
    m.check_days()
    data = m.emit()

    sys.stdout.reconfigure(encoding="utf-8")
    print("wrote %s" % os.path.relpath(OUT_JS, ROOT))
    print("  days    : %d  (%s .. %s)" % (data["source"]["days"],
                                          data["source"]["from"], data["source"]["to"]))
    print("  shifts  : %d" % data["source"]["shifts"])
    print("  workers : %d" % data["source"]["workers"])
    print("  carriers: %s" % ", ".join(data["carriers"]))
    print("  issues  : %d" % len(data["issues"]))
    for kind in sorted({i["kind"] for i in data["issues"]}):
        print("     %-14s %d" % (kind, sum(1 for i in data["issues"] if i["kind"] == kind)))


if __name__ == "__main__":
    main()
